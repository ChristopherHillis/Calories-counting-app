from openpyxl.workbook import workbook
from openpyxl import load_workbook
import os.path
import pandas as pd
from datetime import datetime
import tkinter as tkr

time = datetime.now()
date = datetime.today()
currentDate = date.strftime("%d/%m/%y")
currentTime = time.strftime("%H:%M")

global prompt3

caloriesPath = 'calories.csv'
caloriesData = pd.read_csv(caloriesPath)

if os.path.isfile("test.xlsx") == False:
    testPath = "test.xlsx"
    wb = workbook.Workbook()
    wb.save(filename =testPath)
    book = load_workbook(testPath)
    sheets = book.sheetnames
    Sheet1 = book[sheets[0]]
    
    col = ["Date","Time","Food Item","Type","Calories per 100g","Total"]
    for i in range(1,7):
        Sheet1.cell(row=1,column=i).value=col[i-1]
    book.save(testPath)
else:
    testPath = "test.xlsx"
    book = load_workbook(testPath)
    sheets = book.sheetnames
    Sheet1 = book[sheets[0]]

def select():
    selection = var.get()
    return selection

def submit():
    input = txtInput.get()
    input = input.capitalize()
    selection = select()
    try:
        value = caloriesData.loc[caloriesData['FoodItem'] == input,'Cals_per100grams'].values[0]
        calories = value.split()
        calories = int(calories[0])
        rowValue = [currentDate,currentTime,input,selection,calories]
        maxRow = Sheet1.max_row + 1
        
        for i in range(1,Sheet1.max_column):

            Sheet1.cell(row=maxRow,column=i).value = rowValue[i-1]
            Sheet1.cell(row=Sheet1.max_row,column=Sheet1.max_column).value = "=SUM(E:E)"
            if Sheet1.cell(row=Sheet1.max_row,column=Sheet1.max_column).value == "=SUM(E:E)":
                if Sheet1.cell(row=Sheet1.max_row-1,column=Sheet1.max_column).value != "Total":
                    Sheet1.cell(row=Sheet1.max_row-1,column=Sheet1.max_column).value = " "

        book.save(testPath)   
        prompt3["text"] = "Successfully entered"
            
    except(IndexError):
        prompt3["text"] = "Please try again"
        
tk = tkr.Tk()
tk.title("Calories Tracker Program")

var = tkr.StringVar(value = "Breakfast")

prompt1 = tkr.Label(text="Please enter the food you ate")
prompt2 = tkr.Label(text="Please select what meal the food was a part of")
prompt3 = tkr.Label(text=" ")
button = tkr.Button(master=tk,text="Enter",width = 10,height = 3,command = submit)
r1 = tkr.Radiobutton(master= tk,text = "Breakfast", variable = var,value = "Breakfast",command=select)
r2 = tkr.Radiobutton(master= tk,text = "Lunch", variable = var,value = "Lunch",command=select)
r3 = tkr.Radiobutton(master= tk,text = "Dinner", variable = var,value = "Dinner",command=select)
r4 = tkr.Radiobutton(master= tk,text = "Snack", variable = var,value = "Snack",command=select)
txtInput = tkr.Entry(width=50)
prompt1.pack()
txtInput.pack()
prompt2.pack()
r1.pack()
r2.pack()
r3.pack()
r4.pack()
button.pack()
prompt3.pack()
tk.mainloop()
